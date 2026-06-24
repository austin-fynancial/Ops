import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

# -----------------------------
# Firm Configuration
# -----------------------------
FIRM_NAME = "trillium"
FIRM_DIR = os.path.join(os.path.dirname(__file__), FIRM_NAME)
os.makedirs(FIRM_DIR, exist_ok=True)

# -----------------------------
# Colors
# -----------------------------
DARK_GREEN   = "3A9E68"
LIGHT_GREEN  = "C8E6C4"
LIGHT_YELLOW = "FFF9C4"
LIGHT_RED    = "FFCDD2"
DARK_RED     = "E57373"
LIGHT_BLUE   = "BBDEFB"
PURPLE       = "F2CEEF"

COLOR_RANK = {
    DARK_GREEN:   0,
    LIGHT_GREEN:  1,
    LIGHT_YELLOW: 2,
    LIGHT_RED:    3,
    DARK_RED:     4,
    LIGHT_BLUE:   5,
    PURPLE:       6,
}

# -----------------------------
# Helpers
# -----------------------------
def normalize(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r'\s+', ' ', regex=True)
        .str.replace(r'[^\x20-\x7E]', '', regex=True)
        .str.strip()
    )

def normalize_val(val) -> str:
    s = str(val).strip().lower()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'[^\x20-\x7E]', '', s)
    return s.strip()

def is_household(first_name, last_name=""):
    """Return True if the record appears to be a household (contains & or ' and ')."""
    combined = str(first_name) + " " + str(last_name)
    return bool(re.search(r'&|\band\b', combined, re.IGNORECASE))


def parse_household_members(raw_first, raw_last):
    """
    Parse a household Orion record into a list of (first, last) normalized tuples.

    Handles these patterns:
      1. Standard:      "Aaron & Julie"        | "Mueller"
                        → [("aaron", "mueller"), ("julie", "mueller")]

      2. Trailing &:    "Vanessa Kelly &"      | "Sarah Levenson"
                        → [("vanessa", "kelly"), ("sarah", "levenson")]

      3. Leading & in last: "Andrew Eggum"     | "& Teresa McCarthy-Eggum"
                        → [("andrew", "eggum"), ("teresa", "mccarthy-eggum")]

      4. Full names in first: "Ben Demick & Brandy Blevins" | "Demick"
                        → [("ben", "demick"), ("brandy", "blevins")]
    """
    first = str(raw_first).strip()
    last = str(raw_last).strip()
    splitter = re.compile(r'\s+(?:and|&)\s+|\s*&\s*', re.IGNORECASE)

    members = []

    # Pattern 3: LastName starts with & — "& Teresa McCarthy-Eggum"
    if re.match(r'^&\s*', last):
        person2_full = re.sub(r'^&\s*', '', last).strip()
        person1_full = first
        for full in [person1_full, person2_full]:
            parts = full.split()
            if len(parts) >= 2:
                members.append((normalize_val(parts[0]), normalize_val(" ".join(parts[1:]))))
            elif parts:
                members.append((normalize_val(parts[0]), ""))
        return members

    # Pattern 2: FirstName ends with & — "Vanessa Kelly &"
    if re.search(r'&\s*$', first):
        person1_full = re.sub(r'&\s*$', '', first).strip()
        person2_full = last.strip()
        for full in [person1_full, person2_full]:
            parts = full.split()
            if len(parts) >= 2:
                members.append((normalize_val(parts[0]), normalize_val(" ".join(parts[1:]))))
            elif parts:
                members.append((normalize_val(parts[0]), ""))
        return members

    # Split FirstName on & / and
    parts = [p.strip() for p in splitter.split(first) if p.strip()]

    if len(parts) < 2:
        # No split happened — single name, not really a household
        return [(normalize_val(first), normalize_val(last))]

    # Check if each part already contains a space (full name in FirstName field)
    # Pattern 4: "Ben Demick & Brandy Blevins"
    if all(" " in p for p in parts):
        for full in parts:
            sub = full.split()
            members.append((normalize_val(sub[0]), normalize_val(" ".join(sub[1:]))))
        return members

    # Pattern 1: Standard — each part is just a first name, shared last name
    shared_last = normalize_val(last)
    for p in parts:
        # Part might itself be a full name if it has a space (mixed case like "Andrew Eggum & Teresa")
        sub = p.split()
        if len(sub) >= 2:
            members.append((normalize_val(sub[0]), normalize_val(" ".join(sub[1:]))))
        else:
            members.append((normalize_val(p), shared_last))

    return members

def get_group_key(row):
    # If Orion_Id is shared, that takes priority for grouping
    if pd.notna(row["Orion_Id"]) and row["Orion_Id"] in duplicate_orion_ids:
        return f"orion_{row['Orion_Id']}"
    # Check forced group anchor
    forced = row.get("_forced_group_orion_id")
    if forced is not None and pd.notna(forced) and str(forced).strip() not in ("", "nan"):
        return f"forced_{forced}"
    fyn_orion = row["Fynancial_OrionId"]
    if pd.notna(fyn_orion) and str(fyn_orion).strip() not in ("", "nan"):
        return str(fyn_orion)
    if pd.notna(row["Orion_Id"]) and row["Orion_Id"] in orion_id_to_fyn_orion:
        return orion_id_to_fyn_orion[row["Orion_Id"]]
    return f"single_{row.name}"

# -----------------------------
# Load Files
# -----------------------------
fyn_df = pd.read_csv(os.path.join(FIRM_DIR, f"{FIRM_NAME}_fynancial_contacts.csv"))
orion_df = pd.read_csv(os.path.join(FIRM_DIR, f"{FIRM_NAME}_orion_contacts.csv"))

# Ensure Orion_OrionId and Fynancial_OrionId are strings for matching
orion_df["Orion_OrionId"] = orion_df["Orion_OrionId"].apply(
    lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() not in ("", "nan") else None
)
fyn_df["Fynancial_OrionId"] = fyn_df["Fynancial_OrionId"].apply(
    lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() not in ("", "nan") else None
)

# -----------------------------
# Normalize key columns
# -----------------------------
fyn_df["email_norm"] = normalize(fyn_df["Fynancial_Email"])
fyn_df["name_norm"] = normalize(fyn_df["Fynancial_FirstName"]) + "|" + normalize(fyn_df["Fynancial_LastName"])

# Orion_UserId is the email equivalent
orion_df["email_norm"] = normalize(orion_df["Orion_UserId"])
orion_df["name_norm"] = normalize(orion_df["Orion_FirstName"]) + "|" + normalize(orion_df["Orion_LastName"])
orion_df["is_household"] = orion_df.apply(
    lambda r: is_household(r["Orion_FirstName"], r["Orion_LastName"]), axis=1
)

# -----------------------------
# Output scaffold
# -----------------------------
output_df = fyn_df.copy()
output_df["match_type"] = "No_match"
output_df["Orion_Id"] = None
output_df["Orion_UserId"] = None
output_df["Orion_OrionId"] = None
output_df["Orion_FirstName"] = None
output_df["Orion_LastName"] = None
output_df["fuzzy_score"] = None
output_df["notes"] = None
output_df["_forced_group_orion_id"] = None

# -----------------------------
# 1️⃣ Standard matching (non-household)
# -----------------------------
orion_standard = orion_df[~orion_df["is_household"]]

for idx, row in output_df.iterrows():
    fyn_email = row["email_norm"]
    fyn_name = row["name_norm"]
    fyn_orion_id = row["Fynancial_OrionId"]

    email_and_name = orion_standard.loc[
        (orion_standard["email_norm"] == fyn_email) &
        (orion_standard["name_norm"] == fyn_name)
    ]
    email_match = orion_standard.loc[orion_standard["email_norm"] == fyn_email]
    orion_id_match = orion_standard.loc[
        orion_standard["Orion_OrionId"].notna() &
        (orion_standard["Orion_OrionId"] == fyn_orion_id)
    ] if fyn_orion_id else pd.DataFrame()
    name_match = orion_standard.loc[orion_standard["name_norm"] == fyn_name]

    def orion_id_corroborates(match_row):
        """Return True if the matched Orion row's OrionId agrees with Fynancial_OrionId."""
        return (
            fyn_orion_id is not None and
            pd.notna(match_row["Orion_OrionId"]) and
            str(match_row["Orion_OrionId"]).strip() == str(fyn_orion_id).strip()
        )

    if len(email_and_name) == 1:
        match = email_and_name.iloc[0]
        mt = "Email_And_FullName_And_OrionId" if orion_id_corroborates(match) else "Email_And_FullName"
        output_df.at[idx, "match_type"] = mt
        output_df.at[idx, "Orion_Id"] = match["Orion_Id"]
        output_df.at[idx, "Orion_UserId"] = match["Orion_UserId"]
        output_df.at[idx, "Orion_OrionId"] = match["Orion_OrionId"]
        output_df.at[idx, "Orion_FirstName"] = match["Orion_FirstName"]
        output_df.at[idx, "Orion_LastName"] = match["Orion_LastName"]
        continue

    if len(email_match) == 1:
        match = email_match.iloc[0]
        mt = "Email_And_OrionId" if orion_id_corroborates(match) else "Email"
        output_df.at[idx, "match_type"] = mt
        output_df.at[idx, "Orion_Id"] = match["Orion_Id"]
        output_df.at[idx, "Orion_UserId"] = match["Orion_UserId"]
        output_df.at[idx, "Orion_OrionId"] = match["Orion_OrionId"]
        output_df.at[idx, "Orion_FirstName"] = match["Orion_FirstName"]
        output_df.at[idx, "Orion_LastName"] = match["Orion_LastName"]
        continue

    if len(orion_id_match) == 1:
        match = orion_id_match.iloc[0]
        output_df.at[idx, "match_type"] = "OrionId"
        output_df.at[idx, "Orion_Id"] = match["Orion_Id"]
        output_df.at[idx, "Orion_UserId"] = match["Orion_UserId"]
        output_df.at[idx, "Orion_OrionId"] = match["Orion_OrionId"]
        output_df.at[idx, "Orion_FirstName"] = match["Orion_FirstName"]
        output_df.at[idx, "Orion_LastName"] = match["Orion_LastName"]
        continue

    if len(name_match) == 1:
        match = name_match.iloc[0]
        mt = "FullName_And_OrionId" if orion_id_corroborates(match) else "FullName"
        output_df.at[idx, "match_type"] = mt
        output_df.at[idx, "Orion_Id"] = match["Orion_Id"]
        output_df.at[idx, "Orion_UserId"] = match["Orion_UserId"]
        output_df.at[idx, "Orion_OrionId"] = match["Orion_OrionId"]
        output_df.at[idx, "Orion_FirstName"] = match["Orion_FirstName"]
        output_df.at[idx, "Orion_LastName"] = match["Orion_LastName"]
        continue

# -----------------------------
# 2️⃣ Household matching
# -----------------------------
orion_households = orion_df[orion_df["is_household"]]

for _, orion_row in orion_households.iterrows():
    members = parse_household_members(orion_row["Orion_FirstName"], orion_row["Orion_LastName"])
    member_set = set(members)

    def assign_orion(idx, match_type, _orion_row=orion_row):
        output_df.at[idx, "Orion_Id"] = _orion_row["Orion_Id"]
        output_df.at[idx, "Orion_UserId"] = _orion_row["Orion_UserId"]
        output_df.at[idx, "Orion_OrionId"] = _orion_row["Orion_OrionId"]
        output_df.at[idx, "Orion_FirstName"] = _orion_row["Orion_FirstName"]
        output_df.at[idx, "Orion_LastName"] = _orion_row["Orion_LastName"]
        output_df.at[idx, "match_type"] = match_type

    for idx, fyn_row in output_df.iterrows():
        if fyn_row["match_type"] != "No_match":
            continue

        fyn_first, fyn_last = fyn_row["name_norm"].split("|", 1)

        # Check for perfect raw string match (exact first, last, email, orionid — no normalization)
        fyn_raw_first = str(fyn_row.get("Fynancial_FirstName") or "").strip()
        fyn_raw_last = str(fyn_row.get("Fynancial_LastName") or "").strip()
        orion_raw_first = str(orion_row.get("Orion_FirstName") or "").strip()
        orion_raw_last = str(orion_row.get("Orion_LastName") or "").strip()
        fyn_raw_email = str(fyn_row.get("Fynancial_Email") or "").strip()
        orion_raw_email = str(orion_row.get("Orion_UserId") or "").strip()
        fyn_raw_orion_id = str(fyn_row.get("Fynancial_OrionId") or "").strip()
        orion_raw_orion_id = str(orion_row.get("Orion_OrionId") or "").strip()

        raw_perfect_match = (
            fyn_raw_first == orion_raw_first and
            fyn_raw_last == orion_raw_last and
            fyn_raw_email.lower() == orion_raw_email.lower() and
            fyn_raw_orion_id != "" and orion_raw_orion_id != "" and
            fyn_raw_orion_id == orion_raw_orion_id
        )

        if raw_perfect_match:
            assign_orion(idx, "Household_Perfect")
            continue

        # Check OrionId corroboration for this household row
        fyn_orion_id = fyn_row.get("Fynancial_OrionId")
        orion_orion_id = orion_row.get("Orion_OrionId")
        orion_id_match = (
            fyn_orion_id is not None and
            pd.notna(fyn_orion_id) and str(fyn_orion_id).strip() not in ("", "nan") and
            orion_orion_id is not None and
            pd.notna(orion_orion_id) and
            str(fyn_orion_id).strip() == str(orion_orion_id).strip()
        )

        email_matches = fyn_row["email_norm"] == normalize_val(orion_row["Orion_UserId"])
        name_matches = (fyn_first, fyn_last) in member_set or any(
            m_first and m_first == fyn_first and m_last and m_last == fyn_last
            for (m_first, m_last) in members
        )

        # Assign compound match type based on which signals fired
        if email_matches and name_matches and orion_id_match:
            assign_orion(idx, "Household_Email_And_Name_And_OrionId")
        elif email_matches and orion_id_match:
            assign_orion(idx, "Household_Email_And_OrionId")
        elif name_matches and orion_id_match:
            assign_orion(idx, "Household_Name_And_OrionId")
        elif email_matches and name_matches:
            assign_orion(idx, "Household_Email_And_Name")
        elif email_matches:
            assign_orion(idx, "Household_Email")
        elif name_matches:
            assign_orion(idx, "Household_Name")
        elif orion_id_match:
            assign_orion(idx, "Household_OrionId")

# -----------------------------
# 3️⃣ Fuzzy matching
# -----------------------------
FUZZY_THRESHOLD = 75

already_matched_orion_ids = output_df["Orion_Id"].dropna().unique()
orion_all_unmatched = orion_df[~orion_df["Orion_Id"].isin(already_matched_orion_ids)]

for idx, row in output_df.iterrows():
    if row["match_type"] != "No_match":
        continue

    fyn_first = row["name_norm"].split("|")[0]
    fyn_last = row["name_norm"].split("|")[1]
    fyn_email = row["email_norm"]
    fyn_full = f"{fyn_first} {fyn_last}"

    best_score = 0
    best_match = None

    for _, orion_row in orion_all_unmatched.iterrows():
        orion_full = f"{normalize_val(orion_row['Orion_FirstName'])} {normalize_val(orion_row['Orion_LastName'])}"
        orion_email = orion_row["email_norm"]
        name_score = fuzz.token_sort_ratio(fyn_full, orion_full)
        email_score = fuzz.ratio(fyn_email, orion_email)
        combined_score = round((name_score * 0.6) + (email_score * 0.4), 1)
        if combined_score > best_score:
            best_score = combined_score
            best_match = orion_row

    if best_score >= FUZZY_THRESHOLD and best_match is not None:
        output_df.at[idx, "match_type"] = "Fuzzy_Match"
        output_df.at[idx, "Orion_Id"] = best_match["Orion_Id"]
        output_df.at[idx, "Orion_UserId"] = best_match["Orion_UserId"]
        output_df.at[idx, "Orion_OrionId"] = best_match["Orion_OrionId"]
        output_df.at[idx, "Orion_FirstName"] = best_match["Orion_FirstName"]
        output_df.at[idx, "Orion_LastName"] = best_match["Orion_LastName"]

    output_df.at[idx, "fuzzy_score"] = best_score

# -----------------------------
# 4️⃣ Score all already-matched rows
# -----------------------------
for idx, row in output_df.iterrows():
    if pd.notna(row["fuzzy_score"]):
        continue
    if pd.isna(row["Orion_Id"]):
        continue

    fyn_first = str(row["name_norm"]).split("|")[0] if pd.notna(row["name_norm"]) else ""
    fyn_last = str(row["name_norm"]).split("|")[1] if pd.notna(row["name_norm"]) else ""
    fyn_email = row["email_norm"] if pd.notna(row["email_norm"]) else ""
    fyn_full = f"{fyn_first} {fyn_last}"

    orion_match = orion_df[orion_df["Orion_Id"] == row["Orion_Id"]]
    if len(orion_match) == 0:
        continue
    orion_row = orion_match.iloc[0]
    orion_full = f"{normalize_val(orion_row['Orion_FirstName'])} {normalize_val(orion_row['Orion_LastName'])}"
    orion_email = orion_row["email_norm"]

    name_score = fuzz.token_sort_ratio(fyn_full, orion_full)
    email_score = fuzz.ratio(fyn_email, orion_email)
    output_df.at[idx, "fuzzy_score"] = round((name_score * 0.6) + (email_score * 0.4), 1)

# -----------------------------
# 5️⃣ Detect Household_Partial
# -----------------------------
household_orion_ids = output_df[output_df["match_type"].isin(["Household_Name", "Household_Email", "Household_OrionId", "Household_Email_And_Name", "Household_Email_And_OrionId", "Household_Name_And_OrionId", "Household_Email_And_Name_And_OrionId", "Household_Perfect"])]["Orion_Id"].unique()

for orion_id in household_orion_ids:
    orion_row = orion_df[orion_df["Orion_Id"] == orion_id].iloc[0]
    members = parse_household_members(orion_row["Orion_FirstName"], orion_row["Orion_LastName"])
    all_household_matched = output_df[
        (output_df["Orion_Id"] == orion_id) &
        (output_df["match_type"].isin(["Household_Name", "Household_Email", "Household_OrionId"]))
    ]
    if len(all_household_matched) < len(members):
        name_matched = all_household_matched[all_household_matched["match_type"].isin(["Household_Name", "Household_OrionId", "Household_Email_And_Name", "Household_Email_And_OrionId", "Household_Name_And_OrionId", "Household_Email_And_Name_And_OrionId"])]
        for idx in name_matched.index:
            output_df.at[idx, "match_type"] = "Household_Partial"

# -----------------------------
# 6️⃣ Detect Duplicate_Orion_Id (non-household only)
# -----------------------------
non_household = output_df[
    output_df["Orion_Id"].notna() &
    ~output_df["match_type"].isin(["Household_Name", "Household_Email", "Household_Partial", "Household_OrionId", "Household_Email_And_Name", "Household_Email_And_OrionId", "Household_Name_And_OrionId", "Household_Email_And_Name_And_OrionId", "Household_Perfect"])
]
dup_orion_ids = non_household[non_household.duplicated(subset="Orion_Id", keep=False)]["Orion_Id"].unique()

for idx, row in output_df.iterrows():
    if row["Orion_Id"] in dup_orion_ids and row["match_type"] not in ["Household_Name", "Household_Email", "Household_Partial", "Household_OrionId", "Household_Email_And_Name", "Household_Email_And_OrionId", "Household_Name_And_OrionId", "Household_Email_And_Name_And_OrionId", "Household_Perfect"]:
        output_df.at[idx, "match_type"] = "Duplicate_Orion_Id_Match"

# -----------------------------
# 6b️⃣ Flag missing Fynancial_OrionId
# -----------------------------
for idx, row in output_df.iterrows():
    if row["match_type"] == "Orion_No_Fynancial_Match":
        continue
    fyn_orion = row.get("Fynancial_OrionId")
    is_missing = pd.isna(fyn_orion) or str(fyn_orion).strip() in ("", "nan")
    if is_missing:
        if pd.notna(row["Orion_Id"]):
            output_df.at[idx, "match_type"] = "Matched_No_Fyn_Orion_Id"
        else:
            output_df.at[idx, "match_type"] = "Unmatched_No_Fyn_Orion_Id"

# -----------------------------
# 7️⃣ Shared Fynancial_OrionId — identify shared IDs
# -----------------------------
if "Fynancial_OrionId" in output_df.columns:
    shared_mask = (
        output_df["Fynancial_OrionId"].notna() &
        output_df["Fynancial_OrionId"].astype(str).str.strip().ne("") &
        output_df["Fynancial_OrionId"].astype(str).str.lower().ne("nan")
    )
    orion_id_counts = output_df[shared_mask].groupby("Fynancial_OrionId").size()
    shared_fyn_orion_ids = orion_id_counts[orion_id_counts > 1].index.tolist()

# -----------------------------
# 8️⃣ Unmatched Orion users
# -----------------------------
matched_orion_ids = output_df["Orion_Id"].dropna().unique()
unmatched_orion = orion_df[~orion_df["Orion_Id"].isin(matched_orion_ids)].copy()

unmatched_rows = pd.DataFrame({
    "Fynancial_UserUniqueId": None,
    "Fynancial_Role": None,
    "Fynancial_FirstName": None,
    "Fynancial_LastName": None,
    "Fynancial_Email": None,
    "Fynancial_OrionId": None,
    "email_norm": None,
    "name_norm": None,
    "match_type": "Orion_No_Fynancial_Match",
    "Orion_Id": unmatched_orion["Orion_Id"].values,
    "Orion_UserId": unmatched_orion["Orion_UserId"].values,
    "Orion_OrionId": unmatched_orion["Orion_OrionId"].values,
    "Orion_FirstName": unmatched_orion["Orion_FirstName"].values,
    "Orion_LastName": unmatched_orion["Orion_LastName"].values,
    "fuzzy_score": None,
    "notes": None,
    "_forced_group_orion_id": None,
})

output_df = pd.concat([output_df, unmatched_rows], ignore_index=True)

# -----------------------------
# 8b️⃣ Re-group Orion_No_Fynancial_Match rows whose Orion_UserId matches a Fynancial_Email
# -----------------------------
fyn_email_to_anchor = {}

for _, row in output_df[output_df["match_type"] != "Orion_No_Fynancial_Match"].iterrows():
    fyn_email = row.get("Fynancial_Email")
    if pd.isna(fyn_email) or str(fyn_email).strip() in ("", "nan"):
        continue
    norm_email = normalize_val(fyn_email)
    if not norm_email or norm_email == "nan":
        continue

    fyn_orion = row.get("Fynancial_OrionId")
    orion_id = row.get("Orion_Id")
    fyn_uid = row.get("Fynancial_UserUniqueId")

    if pd.notna(fyn_orion) and str(fyn_orion).strip() not in ("", "nan"):
        anchor = ("fyn_orion", str(fyn_orion))
    elif pd.notna(orion_id):
        anchor = ("orion_id", str(orion_id))
    elif pd.notna(fyn_uid) and str(fyn_uid).strip() not in ("", "nan"):
        anchor = ("fyn_uid", str(fyn_uid))
    else:
        continue

    if norm_email not in fyn_email_to_anchor:
        fyn_email_to_anchor[norm_email] = []
    if anchor not in fyn_email_to_anchor[norm_email]:
        fyn_email_to_anchor[norm_email].append(anchor)

no_fyn_indices = output_df[output_df["match_type"] == "Orion_No_Fynancial_Match"].index.tolist()
indices_to_drop = []
new_rows = []

for idx in no_fyn_indices:
    row = output_df.loc[idx]
    orion_email = row.get("Orion_UserId")
    if pd.isna(orion_email) or str(orion_email).strip() in ("", "nan"):
        continue
    norm_orion_email = normalize_val(orion_email)
    anchors = fyn_email_to_anchor.get(norm_orion_email, [])

    if not anchors:
        continue

    indices_to_drop.append(idx)

    for anchor_type, anchor_val in anchors:
        new_row = row.copy()
        if anchor_type == "fyn_orion":
            new_row["Fynancial_OrionId"] = anchor_val
            new_row["_forced_group_orion_id"] = None
        elif anchor_type == "orion_id":
            new_row["Fynancial_OrionId"] = None
            new_row["_forced_group_orion_id"] = f"orion_id_{anchor_val}"
        else:
            new_row["Fynancial_OrionId"] = None
            new_row["_forced_group_orion_id"] = f"fyn_uid_{anchor_val}"
        new_rows.append(new_row)

if indices_to_drop:
    output_df = output_df.drop(index=indices_to_drop).reset_index(drop=True)
if new_rows:
    output_df = pd.concat([output_df, pd.DataFrame(new_rows)], ignore_index=True)

# -----------------------------
# 9️⃣ Flag shared Orion_Id
# -----------------------------
orion_id_counts = output_df[output_df["Orion_Id"].notna()].groupby("Orion_Id").size()
duplicate_orion_ids = set(orion_id_counts[orion_id_counts > 1].index.tolist())
output_df["dupe_orion_id_found"] = output_df["Orion_Id"].apply(
    lambda x: True if x in duplicate_orion_ids else False
)

# -----------------------------
# Final column order
# -----------------------------
final_columns = [
    "Fynancial_UserUniqueId",
    "Fynancial_Role",
    "Fynancial_FirstName",
    "Fynancial_LastName",
    "Fynancial_Email",
    "Fynancial_OrionId",
    "Orion_Id",
    "Orion_UserId",
    "Orion_OrionId",
    "Orion_FirstName",
    "Orion_LastName",
    "match_type",
    "fuzzy_score",
    "dupe_orion_id_found",
    "notes",
    "_forced_group_orion_id",
]

final_df = output_df[final_columns].copy()
final_df["fuzzy_score"] = pd.to_numeric(final_df["fuzzy_score"], errors="coerce")

# Build Orion_Id -> Fynancial_OrionId lookup
orion_id_to_fyn_orion = {}
for _, row in final_df.iterrows():
    orion_id = row["Orion_Id"]
    fyn_orion = row["Fynancial_OrionId"]
    if pd.notna(orion_id) and pd.notna(fyn_orion) and str(fyn_orion).strip() not in ("", "nan"):
        orion_id_to_fyn_orion[orion_id] = str(fyn_orion)

final_df["_orion_id_sort"] = pd.to_numeric(final_df["Orion_Id"], errors="coerce")

# -----------------------------
# Sort
# -----------------------------
final_df = final_df.sort_values(
    by=["Fynancial_OrionId", "_orion_id_sort", "Fynancial_LastName"],
    na_position="last"
).drop(columns=["_orion_id_sort"]).reset_index(drop=True)

# Post-sort: ensure all rows sharing an Orion_Id are adjacent
processed_orion_ids = set()
new_order = []
used_indices = set()

for i in range(len(final_df)):
    if i in used_indices:
        continue
    orion_id = final_df.at[i, "Orion_Id"]
    if pd.notna(orion_id) and orion_id in duplicate_orion_ids and orion_id not in processed_orion_ids:
        group_indices = final_df[final_df["Orion_Id"] == orion_id].index.tolist()
        new_order.extend(group_indices)
        used_indices.update(group_indices)
        processed_orion_ids.add(orion_id)
    else:
        new_order.append(i)
        used_indices.add(i)

final_df = final_df.loc[new_order].reset_index(drop=True)

# -----------------------------
# Group classification and color assignment
# -----------------------------
final_df["_gk"] = final_df.apply(get_group_key, axis=1)

# Merge group keys: if a Fyn OrionId group contains a member whose Orion_Id is a duplicate,
# redirect all members to the orion_ group key
fyn_orion_to_orion_group = {}
for _, row in final_df.iterrows():
    fyn_orion = row["Fynancial_OrionId"]
    orion_id = row["Orion_Id"]
    if (pd.notna(fyn_orion) and str(fyn_orion).strip() not in ("", "nan") and
            pd.notna(orion_id) and orion_id in duplicate_orion_ids):
        fyn_orion_to_orion_group[str(fyn_orion)] = f"orion_{orion_id}"

def get_merged_gk(row):
    fyn_orion = row["Fynancial_OrionId"]
    if pd.notna(fyn_orion) and str(fyn_orion).strip() not in ("", "nan"):
        fyn_orion_str = str(fyn_orion)
        if fyn_orion_str in fyn_orion_to_orion_group:
            return fyn_orion_to_orion_group[fyn_orion_str]
    return get_group_key(row)

final_df["_gk"] = final_df.apply(get_merged_gk, axis=1)

confident_keys = set()
attention_keys = set()
group_color = {}

for key, grp in final_df.groupby("_gk", sort=False):
    non_orion_types = [mt for mt in grp["match_type"].tolist() if mt != "Orion_No_Fynancial_Match"]
    all_types = grp["match_type"].tolist()
    has_shared = any(grp["dupe_orion_id_found"] == True)
    all_orion_no_fyn = all(mt == "Orion_No_Fynancial_Match" for mt in all_types)

    if all_orion_no_fyn:
        attention_keys.add(key)
        group_color[key] = PURPLE
    elif non_orion_types and all(mt == "Email_And_FullName_And_OrionId" for mt in non_orion_types) and not has_shared:
        confident_keys.add(key)
        group_color[key] = DARK_GREEN
    elif non_orion_types and all(mt in ("Email", "FullName", "Email_And_FullName", "Email_And_FullName_And_OrionId",
                                        "Email_And_OrionId", "FullName_And_OrionId") for mt in non_orion_types) and not has_shared:
        confident_keys.add(key)
        group_color[key] = LIGHT_GREEN
    elif has_shared or all(mt == "No_match" for mt in all_types) or any(mt in ("Matched_No_Fyn_Orion_Id", "Unmatched_No_Fyn_Orion_Id") for mt in all_types):
        fyn_orion_values = grp["Fynancial_OrionId"].apply(
            lambda v: None if pd.isna(v) or str(v).strip() in ("", "nan") else str(v)
        )
        has_some = fyn_orion_values.notna().any()
        has_missing = fyn_orion_values.isna().any()
        unique_ids = fyn_orion_values.dropna().unique()

        if not has_some:
            attention_keys.add(key)
            group_color[key] = LIGHT_BLUE
        elif has_some and (has_missing or len(unique_ids) > 1):
            attention_keys.add(key)
            group_color[key] = DARK_RED
        else:
            attention_keys.add(key)
            group_color[key] = LIGHT_RED
    else:
        attention_keys.add(key)
        group_color[key] = LIGHT_YELLOW

final_df["_row_color"] = final_df["_gk"].map(group_color)

# Force any group containing Household_OrionId to light red (OrionId-only household match, needs review)
for key, grp in final_df.groupby("_gk", sort=False):
    if any(mt == "Household_OrionId" for mt in grp["match_type"].tolist()):
        if group_color.get(key) not in (DARK_RED,):
            group_color[key] = LIGHT_RED
            attention_keys.add(key)
            confident_keys.discard(key)
            for idx in grp.index:
                final_df.at[idx, "_row_color"] = LIGHT_RED

# Force any group containing FullName, Fuzzy_Match, OrionId, or Household_Name to light yellow
REVIEW_TYPES = {"FullName", "FullName_And_OrionId", "Fuzzy_Match", "OrionId", "Household_Name", "Household_Email", "Household_OrionId", "Household_Email_And_Name", "Household_Email_And_OrionId", "Household_Name_And_OrionId", "Household_Email_And_Name_And_OrionId", "Household_Perfect"}
for key, grp in final_df.groupby("_gk", sort=False):
    if any(mt in REVIEW_TYPES for mt in grp["match_type"].tolist()):
        if group_color.get(key) in (DARK_GREEN, LIGHT_GREEN):
            group_color[key] = LIGHT_YELLOW
            attention_keys.add(key)
            confident_keys.discard(key)
            for idx in grp.index:
                final_df.at[idx, "_row_color"] = LIGHT_YELLOW

# Promote single-row Household_Perfect groups to dark green (bottom of dark green)
for key, grp in final_df.groupby("_gk", sort=False):
    non_blank = grp[grp["Fynancial_UserUniqueId"].notna()]
    if (len(non_blank) == 1 and
            all(mt == "Household_Perfect" for mt in grp["match_type"].dropna().tolist())):
        group_color[key] = DARK_GREEN
        confident_keys.add(key)
        attention_keys.discard(key)
        for idx in grp.index:
            final_df.at[idx, "_row_color"] = DARK_GREEN

# Force any group containing an Orion_No_Fynancial_Match row to DARK_RED
orion_no_fyn_group_keys = set()
for key, grp in final_df.groupby("_gk", sort=False):
    if any(mt == "Orion_No_Fynancial_Match" for mt in grp["match_type"].tolist()):
        if not all(mt == "Orion_No_Fynancial_Match" for mt in grp["match_type"].tolist()):
            group_color[key] = DARK_RED
            attention_keys.add(key)
            confident_keys.discard(key)
            orion_no_fyn_group_keys.add(key)
            for idx in grp.index:
                final_df.at[idx, "_row_color"] = DARK_RED

color_lookup = {}
for _, row in final_df.iterrows():
    uid = row.get("Fynancial_UserUniqueId")
    color = row.get("_row_color")
    if pd.notna(uid) and str(uid).strip() not in ("", "nan"):
        color_lookup[uid] = color
    orion_id = row.get("Orion_Id")
    if pd.notna(orion_id):
        color_lookup.setdefault(f"_orion_{orion_id}", color)

final_df = final_df.drop(columns=["_gk", "_row_color"])

# -----------------------------
# Split into confident and attention sections
# -----------------------------
def row_in_confident(row):
    key = get_merged_gk(row)
    return key in confident_keys and key not in orion_no_fyn_group_keys

df_confident = final_df[final_df.apply(row_in_confident, axis=1)].copy()
df_attention = final_df[~final_df.apply(row_in_confident, axis=1)].copy()

def group_by_orion_id(df):
    processed = set()
    new_order = []
    used = set()
    for i in range(len(df)):
        if i in used:
            continue
        orion_id = df.iloc[i]["Orion_Id"]
        if pd.notna(orion_id) and orion_id in duplicate_orion_ids and orion_id not in processed:
            group_indices = df[df["Orion_Id"] == orion_id].index.tolist()
            pos_indices = [df.index.get_loc(gi) for gi in group_indices]
            new_order.extend(pos_indices)
            used.update(pos_indices)
            processed.add(orion_id)
        else:
            new_order.append(i)
            used.add(i)
    return df.iloc[new_order].reset_index(drop=True)

df_confident = group_by_orion_id(df_confident)
df_attention = group_by_orion_id(df_attention)

# Move any confident rows whose duplicate partner is in attention
dupe_orion_ids_in_attention = set(df_attention[df_attention["dupe_orion_id_found"] == True]["Orion_Id"].dropna().unique())
rows_to_move = df_confident[df_confident["Orion_Id"].isin(dupe_orion_ids_in_attention)]
if len(rows_to_move) > 0:
    df_confident = df_confident[~df_confident["Orion_Id"].isin(dupe_orion_ids_in_attention)].reset_index(drop=True)
    df_attention = pd.concat([rows_to_move, df_attention], ignore_index=True)
    df_attention = group_by_orion_id(df_attention)

# -----------------------------
# Sort sections by color rank
# -----------------------------
def sort_by_color(df):
    df = df.copy()
    gk_series = df.apply(get_merged_gk, axis=1)
    group_sizes = gk_series.map(gk_series.value_counts())

    gk_has_match = {}
    for key, grp in df.groupby(gk_series, sort=False):
        gk_has_match[key] = 0 if any(pd.notna(v) for v in grp["Orion_Id"]) else 1
    has_match_series = gk_series.map(gk_has_match)

    gk_has_nomatch = {}
    for key, grp in df.groupby(gk_series, sort=False):
        gk_has_nomatch[key] = 1 if any(mt == "No_match" for mt in grp["match_type"]) else 0
    has_nomatch_series = gk_series.map(gk_has_nomatch)

    def get_row_color(r):
        uid = r.get("Fynancial_UserUniqueId")
        if pd.notna(uid) and str(uid).strip() not in ("", "nan"):
            return color_lookup.get(uid)
        orion_id = r.get("Orion_Id")
        if pd.notna(orion_id):
            return color_lookup.get(f"_orion_{orion_id}")
        return None

    df["_cr"] = df.apply(lambda r: COLOR_RANK.get(get_row_color(r), 99), axis=1)
    df["_group_size"] = group_sizes.values
    df["_has_match"] = has_match_series.values
    df["_has_nomatch"] = has_nomatch_series.values
    df["_orig"] = range(len(df))
    df["_has_orion_no_fyn"] = df.apply(
        lambda r: 1 if get_merged_gk(r) in orion_no_fyn_group_keys else 0, axis=1
    )
    df["_match_sort"] = df.apply(
        lambda r: r["_has_match"] if get_row_color(r) in (LIGHT_RED, DARK_RED) else 0,
        axis=1
    )
    df["_yellow_nomatch_sort"] = df.apply(
        lambda r: r["_has_nomatch"] if (
            group_sizes[r.name] > 1 and get_row_color(r) == LIGHT_YELLOW
        ) else 0,
        axis=1
    )
    match_type_order = {
        "Email_And_FullName_And_OrionId": 0, "Email_And_FullName": 1,
        "Email_And_OrionId": 2, "Email": 3,
        "FullName_And_OrionId": 4, "OrionId": 5, "FullName": 6,
        "Fuzzy_Match": 7,
        "Household_Perfect": 8, "Household_Email_And_Name_And_OrionId": 9, "Household_Email_And_Name": 10,
        "Household_Email_And_OrionId": 11, "Household_Email": 12,
        "Household_Name_And_OrionId": 13, "Household_Name": 14,
        "Household_OrionId": 15, "Household_Partial": 16,
        "Duplicate_Orion_Id_Match": 17,
        "Matched_No_Fyn_Orion_Id": 18, "Unmatched_No_Fyn_Orion_Id": 19,
        "No_match": 20, "Orion_No_Fynancial_Match": 21,
    }
    df["_fuzzy_sort"] = df.apply(
        lambda r: -r["fuzzy_score"] if (
            group_sizes[r.name] == 1 and
            get_row_color(r) in (LIGHT_GREEN, LIGHT_YELLOW) and
            pd.notna(r["fuzzy_score"])
        ) else 0,
        axis=1
    )
    df["_mt_sort"] = df.apply(
        lambda r: match_type_order.get(r["match_type"], 99) if (
            group_sizes[r.name] == 1 and
            get_row_color(r) in (LIGHT_GREEN, LIGHT_YELLOW)
        ) else 0,
        axis=1
    )
    df = df.sort_values(
        ["_cr", "_has_orion_no_fyn", "_group_size", "_match_sort", "_yellow_nomatch_sort", "_mt_sort", "_fuzzy_sort", "_orig"],
        ascending=[True, True, False, True, True, True, True, True]
    ).drop(columns=["_cr", "_has_orion_no_fyn", "_group_size", "_has_match", "_has_nomatch",
                    "_match_sort", "_yellow_nomatch_sort", "_mt_sort", "_fuzzy_sort", "_orig"])
    return df.reset_index(drop=True)

df_confident = sort_by_color(df_confident)
df_attention = sort_by_color(df_attention)

# -----------------------------
# Insert blank rows between groups
# -----------------------------
def insert_blanks(df):
    blank = pd.Series({col: None for col in df.columns})
    rows = []
    prev_key = "SENTINEL"
    for i, row in df.iterrows():
        curr_key = get_merged_gk(row)
        if i > 0 and curr_key != prev_key:
            rows.append(blank)
        rows.append(row)
        prev_key = curr_key
    return pd.DataFrame(rows).reset_index(drop=True)

excel_confident = insert_blanks(df_confident)
excel_attention = insert_blanks(df_attention)
section_blank = pd.DataFrame([{col: None for col in final_df.columns}] * 2)
final_df_excel = pd.concat([excel_confident, section_blank, excel_attention], ignore_index=True)

final_df_excel = final_df_excel.drop(columns=["_forced_group_orion_id"], errors="ignore")
final_df = final_df.drop(columns=["_forced_group_orion_id"], errors="ignore")

# -----------------------------
# Write Excel
# -----------------------------
output_filename = os.path.join(FIRM_DIR, f"{FIRM_NAME}_fynancial_orion_mapping.xlsx")
final_df_excel.to_excel(output_filename, index=False)

wb = load_workbook(output_filename)
ws = wb.active

# Auto-size columns
for col_idx, col in enumerate(final_df_excel.columns, start=1):
    max_len = max(len(str(col)), final_df_excel[col].astype(str).str.len().max())
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

# Column dividers
divider_side = Side(style="medium")
divider_cols = [
    final_df_excel.columns.get_loc("Fynancial_OrionId") + 1,
    final_df_excel.columns.get_loc("Orion_LastName") + 1,
]

# Bold header
for cell in ws[1]:
    cell.font = Font(bold=True)

# Color rows
match_type_col = final_df_excel.columns.get_loc("match_type") + 1
shared_col = final_df_excel.columns.get_loc("dupe_orion_id_found") + 1
uid_col = final_df_excel.columns.get_loc("Fynancial_UserUniqueId") + 1
orion_id_col = final_df_excel.columns.get_loc("Orion_Id") + 1

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    uid = row[uid_col - 1].value
    orion_id = row[orion_id_col - 1].value

    if uid and uid in color_lookup:
        hex_color = color_lookup[uid]
    elif orion_id and f"_orion_{orion_id}" in color_lookup:
        hex_color = color_lookup[f"_orion_{orion_id}"]
    else:
        hex_color = None

    for cell in row:
        if hex_color:
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        if cell.column in divider_cols:
            existing = cell.border
            cell.border = Border(
                left=existing.left,
                right=divider_side,
                top=existing.top,
                bottom=existing.bottom
            )

    shared_cell = row[shared_col - 1]
    if shared_cell.value is True:
        row_color = color_lookup.get(uid) or color_lookup.get(f"_orion_{orion_id}", LIGHT_RED)
        if row_color == DARK_RED:
            dupe_color = DARK_RED
        elif row_color == LIGHT_BLUE:
            dupe_color = LIGHT_BLUE
        else:
            dupe_color = LIGHT_RED
        shared_cell.fill = PatternFill(start_color=dupe_color, end_color=dupe_color, fill_type="solid")

# Merge notes cells per group
notes_col = final_df_excel.columns.get_loc("notes") + 1

def get_excel_group_key(row):
    fyn_orion = row.get("Fynancial_OrionId")
    orion_id = row.get("Orion_Id")
    uid = row.get("Fynancial_UserUniqueId")
    if pd.notna(fyn_orion) and str(fyn_orion).strip() not in ("", "nan"):
        return str(fyn_orion)
    if pd.notna(orion_id) and orion_id in orion_id_to_fyn_orion:
        return orion_id_to_fyn_orion[orion_id]
    if pd.notna(orion_id) and orion_id in duplicate_orion_ids:
        return f"orion_{orion_id}"
    return f"single_{uid}"

rows_with_keys = []
for i, row in final_df_excel.iterrows():
    if all(pd.isna(v) or v is None for v in row.values):
        rows_with_keys.append((i, None))
    else:
        rows_with_keys.append((i, get_excel_group_key(row)))

i = 0
while i < len(rows_with_keys):
    df_idx, key = rows_with_keys[i]
    if key is None:
        i += 1
        continue
    j = i + 1
    while j < len(rows_with_keys) and rows_with_keys[j][1] == key:
        j += 1
    excel_start = df_idx + 2
    excel_end = rows_with_keys[j - 1][0] + 2
    if excel_end > excel_start:
        ws.merge_cells(
            start_row=excel_start, start_column=notes_col,
            end_row=excel_end, end_column=notes_col
        )
    merged_cell = ws.cell(row=excel_start, column=notes_col)
    merged_cell.border = Border(
        top=Side(style="thin"), bottom=Side(style="thin"),
        left=Side(style="thin"), right=Side(style="thin")
    )
    i = j

wb.save(output_filename)

# -----------------------------
# Draw thick border boxes around groups of 2+ rows
# -----------------------------
wb = load_workbook(output_filename)
ws = wb.active

thick = Side(style="medium", color="000000")

row_group_keys = []
for i, row in final_df_excel.iterrows():
    if all(pd.isna(v) or v is None for v in row.values):
        row_group_keys.append((i + 2, None))
    else:
        row_group_keys.append((i + 2, get_excel_group_key(row)))

i = 0
while i < len(row_group_keys):
    excel_row, key = row_group_keys[i]
    if key is None:
        i += 1
        continue
    j = i + 1
    while j < len(row_group_keys) and row_group_keys[j][1] == key:
        j += 1
    group_excel_rows = [r for r, k in row_group_keys[i:j] if k is not None]
    if len(group_excel_rows) > 1:
        first_row = min(group_excel_rows)
        last_row = max(group_excel_rows)
        max_col = len(final_df_excel.columns)
        for excel_row in group_excel_rows:
            for col in range(1, max_col + 1):
                cell = ws.cell(row=excel_row, column=col)
                top = thick if excel_row == first_row else cell.border.top
                bottom = thick if excel_row == last_row else cell.border.bottom
                left = thick if col == 1 else cell.border.left
                right = thick if col == max_col else cell.border.right
                if col in divider_cols:
                    right = thick if col == max_col else Side(style="medium")
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)
    i = j

wb.save(output_filename)

# -----------------------------
# Also write CSV
# -----------------------------
csv_filename = os.path.join(FIRM_DIR, f"{FIRM_NAME}_fynancial_orion_mapping.csv")
final_df.to_csv(csv_filename, index=False)

fyn_row_count = len(fyn_df)
orion_unmatched_count = len(final_df[final_df["match_type"] == "Orion_No_Fynancial_Match"])
assert len(final_df) == fyn_row_count + orion_unmatched_count, "Row count mismatch!"

print(f"✅ Excel generated: {output_filename}")
print(f"✅ CSV generated:   {csv_filename}")
print(f"\nSummary:")
print(f"  Total records:              {len(final_df)}")
print(f"  Email And FullName And OrionId: {len(final_df[final_df['match_type'] == 'Email_And_FullName_And_OrionId'])}")
print(f"  Email And FullName:         {len(final_df[final_df['match_type'] == 'Email_And_FullName'])}")
print(f"  Email And OrionId:          {len(final_df[final_df['match_type'] == 'Email_And_OrionId'])}")
print(f"  Email:                      {len(final_df[final_df['match_type'] == 'Email'])}")
print(f"  FullName And OrionId:       {len(final_df[final_df['match_type'] == 'FullName_And_OrionId'])}")
print(f"  OrionId:                    {len(final_df[final_df['match_type'] == 'OrionId'])}")
print(f"  FullName:                   {len(final_df[final_df['match_type'] == 'FullName'])}")
print(f"  Fuzzy Match:                {len(final_df[final_df['match_type'] == 'Fuzzy_Match'])}")
print(f"  Household Perfect:          {len(final_df[final_df['match_type'] == 'Household_Perfect'])}")
print(f"  Household Email+Name+OrionId: {len(final_df[final_df['match_type'] == 'Household_Email_And_Name_And_OrionId'])}")
print(f"  Household Email+Name:       {len(final_df[final_df['match_type'] == 'Household_Email_And_Name'])}")
print(f"  Household Email+OrionId:    {len(final_df[final_df['match_type'] == 'Household_Email_And_OrionId'])}")
print(f"  Household Email:            {len(final_df[final_df['match_type'] == 'Household_Email'])}")
print(f"  Household Name+OrionId:     {len(final_df[final_df['match_type'] == 'Household_Name_And_OrionId'])}")
print(f"  Household Name:             {len(final_df[final_df['match_type'] == 'Household_Name'])}")
print(f"  Household OrionId:          {len(final_df[final_df['match_type'] == 'Household_OrionId'])}")
print(f"  Household Partial:          {len(final_df[final_df['match_type'] == 'Household_Partial'])}")
print(f"  Duplicate Orion Id:         {len(final_df[final_df['match_type'] == 'Duplicate_Orion_Id_Match'])}")
print(f"  Matched No Fyn Orion Id:    {len(final_df[final_df['match_type'] == 'Matched_No_Fyn_Orion_Id'])}")
print(f"  Unmatched No Fyn Orion Id:  {len(final_df[final_df['match_type'] == 'Unmatched_No_Fyn_Orion_Id'])}")
print(f"  No match:                   {len(final_df[final_df['match_type'] == 'No_match'])}")
print(f"  Orion no Fynancial:         {len(final_df[final_df['match_type'] == 'Orion_No_Fynancial_Match'])}")