import pandas as pd
import json
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

# -----------------------------
# Firm Configuration
# -----------------------------
FIRM_NAME = "legacy_financial"
FIRM_DIR = os.path.join(os.path.dirname(__file__), FIRM_NAME)
os.makedirs(FIRM_DIR, exist_ok=True)

# -----------------------------
# Colors
# -----------------------------
DARK_GREEN   = "1E8449"
LIGHT_GREEN  = "C8E6C4"
LIGHT_YELLOW = "FFF9C4"
LIGHT_RED    = "FFCDD2"
DARK_RED     = "E57373"
LIGHT_BLUE   = "BBDEFB"
PURPLE       = "F2CEEF"

COLOR_RANK = {
    DARK_GREEN:   0,
    LIGHT_GREEN:  1,
    LIGHT_YELLOW: 2,
    LIGHT_RED:    3,
    DARK_RED:     4,
    LIGHT_BLUE:   5,
    PURPLE:       6,
}

# -----------------------------
# Helpers
# -----------------------------
def normalize(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r'\s+', ' ', regex=True)
        .str.replace(r'[^\x20-\x7E]', '', regex=True)
        .str.strip()
    )

def normalize_val(val) -> str:
    s = str(val).strip().lower()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'[^\x20-\x7E]', '', s)
    return s.strip()

def get_group_key(row):
    # If BD_UserID is shared, that takes priority for grouping
    if pd.notna(row["BD_UserID"]) and row["BD_UserID"] in duplicate_bd_userids:
        return f"bd_{row['BD_UserID']}"
    # Check forced group anchor (for re-grouped BD_No_Fynancial_Match rows with no Fyn BD ID)
    forced = row.get("_forced_group_bd_uid")
    if forced is not None and pd.notna(forced) and str(forced).strip() not in ("", "nan"):
        return f"forced_{forced}"
    fyn_bd = row["Fynancial_BlackDiamondId"]
    if pd.notna(fyn_bd) and str(fyn_bd).strip() not in ("", "nan"):
        return str(fyn_bd)
    if pd.notna(row["BD_UserID"]) and row["BD_UserID"] in bd_userid_to_fyn_bd:
        return bd_userid_to_fyn_bd[row["BD_UserID"]]
    return f"single_{row.name}"

# -----------------------------
# Load Files
# -----------------------------
fyn_df = pd.read_csv(os.path.join(FIRM_DIR, f"{FIRM_NAME}_fynancial_contacts.csv"))

with open(os.path.join(FIRM_DIR, f"{FIRM_NAME}_blackdiamond_contacts.json"), "r") as f:
    bd_raw = json.load(f)

bd_df = pd.DataFrame(bd_raw)

# -----------------------------
# Normalize key columns
# -----------------------------
fyn_df["email_norm"] = normalize(fyn_df["Fynancial_Email"])
fyn_df["name_norm"] = normalize(fyn_df["Fynancial_FirstName"]) + "|" + normalize(fyn_df["Fynancial_LastName"])

bd_df["email_norm"] = normalize(bd_df["Email"])
bd_df["name_norm"] = normalize(bd_df["FirstName"]) + "|" + normalize(bd_df["LastName"])

def is_household(first_name):
    return bool(re.search(r'&|\band\b', str(first_name), re.IGNORECASE))

bd_df["is_household"] = bd_df["FirstName"].apply(is_household)

# -----------------------------
# Output scaffold
# -----------------------------
output_df = fyn_df.copy()
output_df["match_type"] = "No_match"
output_df["BD_UserID"] = None
output_df["BD_Email"] = None
output_df["BD_FirstName"] = None
output_df["BD_LastName"] = None
output_df["fuzzy_score"] = None
output_df["notes"] = None
output_df["_forced_group_bd_uid"] = None

# -----------------------------
# 1️⃣ Standard matching
# -----------------------------
bd_standard = bd_df[~bd_df["is_household"]]

for idx, row in output_df.iterrows():
    fyn_email = row["email_norm"]
    fyn_name = row["name_norm"]

    email_match = bd_standard.loc[bd_standard["email_norm"] == fyn_email]
    name_match = bd_standard.loc[bd_standard["name_norm"] == fyn_name]
    email_and_name = bd_standard.loc[
        (bd_standard["email_norm"] == fyn_email) &
        (bd_standard["name_norm"] == fyn_name)
    ]

    if len(email_and_name) == 1:
        match = email_and_name.iloc[0]
        output_df.at[idx, "match_type"] = "Email_And_FullName"
        output_df.at[idx, "BD_UserID"] = match["UserID"]
        output_df.at[idx, "BD_Email"] = match["Email"]
        output_df.at[idx, "BD_FirstName"] = match["FirstName"]
        output_df.at[idx, "BD_LastName"] = match["LastName"]
        continue

    if len(email_match) == 1:
        match = email_match.iloc[0]
        output_df.at[idx, "match_type"] = "Email"
        output_df.at[idx, "BD_UserID"] = match["UserID"]
        output_df.at[idx, "BD_Email"] = match["Email"]
        output_df.at[idx, "BD_FirstName"] = match["FirstName"]
        output_df.at[idx, "BD_LastName"] = match["LastName"]
        continue

    if len(name_match) == 1:
        match = name_match.iloc[0]
        output_df.at[idx, "match_type"] = "FullName"
        output_df.at[idx, "BD_UserID"] = match["UserID"]
        output_df.at[idx, "BD_Email"] = match["Email"]
        output_df.at[idx, "BD_FirstName"] = match["FirstName"]
        output_df.at[idx, "BD_LastName"] = match["LastName"]
        continue

# -----------------------------
# 2️⃣ Household matching
# -----------------------------
bd_households = bd_df[bd_df["is_household"]]

for _, bd_row in bd_households.iterrows():
    raw_first = str(bd_row["FirstName"])
    last = normalize_val(bd_row["LastName"])
    parts = [p.strip() for p in re.split(r'\s+and\s+|&', raw_first, flags=re.IGNORECASE) if p.strip()]

    for idx, fyn_row in output_df.iterrows():
        if fyn_row["match_type"] != "No_match":
            continue
        fyn_first, fyn_last = fyn_row["name_norm"].split("|")

        if fyn_row["email_norm"] == normalize_val(bd_row["Email"]):
            output_df.at[idx, "BD_UserID"] = bd_row["UserID"]
            output_df.at[idx, "BD_Email"] = bd_row["Email"]
            output_df.at[idx, "BD_FirstName"] = bd_row["FirstName"]
            output_df.at[idx, "BD_LastName"] = bd_row["LastName"]
            output_df.at[idx, "match_type"] = "Household_Email"
            continue

        if fyn_last != last:
            continue
        if fyn_first in [normalize_val(p) for p in parts]:
            output_df.at[idx, "BD_UserID"] = bd_row["UserID"]
            output_df.at[idx, "BD_Email"] = bd_row["Email"]
            output_df.at[idx, "BD_FirstName"] = bd_row["FirstName"]
            output_df.at[idx, "BD_LastName"] = bd_row["LastName"]
            output_df.at[idx, "match_type"] = "Household_Name"

# -----------------------------
# 3️⃣ Fuzzy matching
# -----------------------------
FUZZY_THRESHOLD = 75

already_matched_bd_ids = output_df["BD_UserID"].dropna().unique()
bd_all_unmatched = bd_df[~bd_df["UserID"].isin(already_matched_bd_ids)]

for idx, row in output_df.iterrows():
    if row["match_type"] != "No_match":
        continue

    fyn_first = row["name_norm"].split("|")[0]
    fyn_last = row["name_norm"].split("|")[1]
    fyn_email = row["email_norm"]
    fyn_full = f"{fyn_first} {fyn_last}"

    best_score = 0
    best_match = None

    for _, bd_row in bd_all_unmatched.iterrows():
        bd_full = f"{normalize_val(bd_row['FirstName'])} {normalize_val(bd_row['LastName'])}"
        bd_email = bd_row["email_norm"]
        name_score = fuzz.token_sort_ratio(fyn_full, bd_full)
        email_score = fuzz.ratio(fyn_email, bd_email)
        combined_score = round((name_score * 0.6) + (email_score * 0.4), 1)
        if combined_score > best_score:
            best_score = combined_score
            best_match = bd_row

    if best_score >= FUZZY_THRESHOLD and best_match is not None:
        output_df.at[idx, "match_type"] = "Fuzzy_Match"
        output_df.at[idx, "BD_UserID"] = best_match["UserID"]
        output_df.at[idx, "BD_Email"] = best_match["Email"]
        output_df.at[idx, "BD_FirstName"] = best_match["FirstName"]
        output_df.at[idx, "BD_LastName"] = best_match["LastName"]

    output_df.at[idx, "fuzzy_score"] = best_score

# -----------------------------
# 4️⃣ Score all already-matched rows
# -----------------------------
for idx, row in output_df.iterrows():
    if pd.notna(row["fuzzy_score"]):
        continue
    if pd.isna(row["BD_UserID"]):
        continue

    fyn_first = str(row["name_norm"]).split("|")[0] if pd.notna(row["name_norm"]) else ""
    fyn_last = str(row["name_norm"]).split("|")[1] if pd.notna(row["name_norm"]) else ""
    fyn_email = row["email_norm"] if pd.notna(row["email_norm"]) else ""
    fyn_full = f"{fyn_first} {fyn_last}"

    bd_match = bd_df[bd_df["UserID"] == row["BD_UserID"]]
    if len(bd_match) == 0:
        continue
    bd_row = bd_match.iloc[0]
    bd_full = f"{normalize_val(bd_row['FirstName'])} {normalize_val(bd_row['LastName'])}"
    bd_email = bd_row["email_norm"]

    name_score = fuzz.token_sort_ratio(fyn_full, bd_full)
    email_score = fuzz.ratio(fyn_email, bd_email)
    output_df.at[idx, "fuzzy_score"] = round((name_score * 0.6) + (email_score * 0.4), 1)

# -----------------------------
# 5️⃣ Detect Household_Partial
# -----------------------------
household_bd_ids = output_df[output_df["match_type"].isin(["Household_Name", "Household_Email"])]["BD_UserID"].unique()

for bd_id in household_bd_ids:
    bd_row = bd_df[bd_df["UserID"] == bd_id].iloc[0]
    raw_first = str(bd_row["FirstName"])
    parts = [p.strip() for p in re.split(r'\s+and\s+|&', raw_first, flags=re.IGNORECASE) if p.strip()]
    all_household_matched = output_df[
        (output_df["BD_UserID"] == bd_id) &
        (output_df["match_type"].isin(["Household_Name", "Household_Email"]))
    ]
    if len(all_household_matched) < len(parts):
        name_matched = all_household_matched[all_household_matched["match_type"] == "Household_Name"]
        for idx in name_matched.index:
            output_df.at[idx, "match_type"] = "Household_Partial"

# -----------------------------
# 6️⃣ Detect Duplicate_BD_UserID (non-household only)
# -----------------------------
non_household = output_df[
    output_df["BD_UserID"].notna() &
    ~output_df["match_type"].isin(["Household_Name", "Household_Email", "Household_Partial"])
]
dup_bd_ids = non_household[non_household.duplicated(subset="BD_UserID", keep=False)]["BD_UserID"].unique()

for idx, row in output_df.iterrows():
    if row["BD_UserID"] in dup_bd_ids and row["match_type"] not in ["Household_Name", "Household_Email", "Household_Partial"]:
        output_df.at[idx, "match_type"] = "Duplicate_BD_UserID_Match"

# -----------------------------
# 6b️⃣ Flag missing Fynancial_BlackDiamondId
# -----------------------------
for idx, row in output_df.iterrows():
    if row["match_type"] == "BD_No_Fynancial_Match":
        continue
    fyn_bd = row.get("Fynancial_BlackDiamondId")
    is_missing = pd.isna(fyn_bd) or str(fyn_bd).strip() in ("", "nan")
    if is_missing:
        if pd.notna(row["BD_UserID"]):
            output_df.at[idx, "match_type"] = "Matched_No_Fyn_BD_ID"
        else:
            output_df.at[idx, "match_type"] = "Unmatched_No_Fyn_BD_ID"

# -----------------------------
# 7️⃣ Shared Fynancial_BlackDiamondId — identify shared IDs only
# -----------------------------
if "Fynancial_BlackDiamondId" in output_df.columns:
    shared_mask = (
        output_df["Fynancial_BlackDiamondId"].notna() &
        output_df["Fynancial_BlackDiamondId"].astype(str).str.strip().ne("") &
        output_df["Fynancial_BlackDiamondId"].astype(str).str.lower().ne("nan")
    )
    bd_id_counts = output_df[shared_mask].groupby("Fynancial_BlackDiamondId").size()
    shared_bd_ids = bd_id_counts[bd_id_counts > 1].index.tolist()

# -----------------------------
# 8️⃣ Unmatched BD users
# -----------------------------
matched_bd_ids = output_df["BD_UserID"].dropna().unique()
unmatched_bd = bd_df[~bd_df["UserID"].isin(matched_bd_ids)].copy()

unmatched_rows = pd.DataFrame({
    "Fynancial_UserUniqueId": None,
    "Fynancial_Role": None,
    "Fynancial_FirstName": None,
    "Fynancial_LastName": None,
    "Fynancial_Email": None,
    "Fynancial_BlackDiamondId": None,
    "email_norm": None,
    "name_norm": None,
    "match_type": "BD_No_Fynancial_Match",
    "BD_UserID": unmatched_bd["UserID"].values,
    "BD_Email": unmatched_bd["Email"].values,
    "BD_FirstName": unmatched_bd["FirstName"].values,
    "BD_LastName": unmatched_bd["LastName"].values,
    "fuzzy_score": None,
    "notes": None,
    "_forced_group_bd_uid": None,
})

output_df = pd.concat([output_df, unmatched_rows], ignore_index=True)

# -----------------------------
# 8b️⃣ Re-group BD_No_Fynancial_Match rows whose BD_Email matches a Fynancial_Email
# -----------------------------
# Build lookup: normalized Fynancial_Email -> group anchor
# Anchor preference: Fynancial_BlackDiamondId > BD_UserID > Fynancial_UserUniqueId sentinel
fyn_email_to_anchor = {}  # norm_email -> list of anchors

for _, row in output_df[output_df["match_type"] != "BD_No_Fynancial_Match"].iterrows():
    fyn_email = row.get("Fynancial_Email")
    if pd.isna(fyn_email) or str(fyn_email).strip() in ("", "nan"):
        continue
    norm_email = normalize_val(fyn_email)
    if not norm_email or norm_email == "nan":
        continue

    fyn_bd = row.get("Fynancial_BlackDiamondId")
    bd_uid = row.get("BD_UserID")
    fyn_uid = row.get("Fynancial_UserUniqueId")

    if pd.notna(fyn_bd) and str(fyn_bd).strip() not in ("", "nan"):
        anchor = ("fyn_bd", str(fyn_bd))
    elif pd.notna(bd_uid):
        anchor = ("bd_uid", str(bd_uid))
    elif pd.notna(fyn_uid) and str(fyn_uid).strip() not in ("", "nan"):
        anchor = ("fyn_uid", str(fyn_uid))
    else:
        continue

    if norm_email not in fyn_email_to_anchor:
        fyn_email_to_anchor[norm_email] = []
    if anchor not in fyn_email_to_anchor[norm_email]:
        fyn_email_to_anchor[norm_email].append(anchor)

# For each BD_No_Fynancial_Match row, check if BD_Email matches any Fynancial_Email
no_fyn_indices = output_df[output_df["match_type"] == "BD_No_Fynancial_Match"].index.tolist()
indices_to_drop = []
new_rows = []

for idx in no_fyn_indices:
    row = output_df.loc[idx]
    bd_email = row.get("BD_Email")
    if pd.isna(bd_email) or str(bd_email).strip() in ("", "nan"):
        continue
    norm_bd_email = normalize_val(bd_email)
    anchors = fyn_email_to_anchor.get(norm_bd_email, [])

    if not anchors:
        continue

    # Drop the original row; replace with one copy per matching group
    indices_to_drop.append(idx)

    for anchor_type, anchor_val in anchors:
        new_row = row.copy()
        if anchor_type == "fyn_bd":
            # Assign the group's Fynancial_BlackDiamondId so grouping logic picks it up naturally
            new_row["Fynancial_BlackDiamondId"] = anchor_val
            new_row["_forced_group_bd_uid"] = None
        elif anchor_type == "bd_uid":
            # Group has no Fyn BD ID — use forced anchor keyed by BD_UserID
            new_row["Fynancial_BlackDiamondId"] = None
            new_row["_forced_group_bd_uid"] = f"bd_uid_{anchor_val}"
        else:
            # fyn_uid fallback — use forced anchor keyed by Fynancial_UserUniqueId
            new_row["Fynancial_BlackDiamondId"] = None
            new_row["_forced_group_bd_uid"] = f"fyn_uid_{anchor_val}"
        new_rows.append(new_row)

if indices_to_drop:
    output_df = output_df.drop(index=indices_to_drop).reset_index(drop=True)
if new_rows:
    output_df = pd.concat([output_df, pd.DataFrame(new_rows)], ignore_index=True)

# -----------------------------
# 9️⃣ Flag shared BD UserIDs
# -----------------------------
bd_userid_counts = output_df[output_df["BD_UserID"].notna()].groupby("BD_UserID").size()
duplicate_bd_userids = set(bd_userid_counts[bd_userid_counts > 1].index.tolist())
output_df["dupe_bd_userid_found"] = output_df["BD_UserID"].apply(
    lambda x: True if x in duplicate_bd_userids else False
)

# -----------------------------
# Final column order
# -----------------------------
final_columns = [
    "Fynancial_UserUniqueId",
    "Fynancial_Role",
    "Fynancial_FirstName",
    "Fynancial_LastName",
    "Fynancial_Email",
    "Fynancial_BlackDiamondId",
    "BD_UserID",
    "BD_Email",
    "BD_FirstName",
    "BD_LastName",
    "match_type",
    "fuzzy_score",
    "dupe_bd_userid_found",
    "notes",
    "_forced_group_bd_uid",
]

final_df = output_df[final_columns].copy()
final_df["fuzzy_score"] = pd.to_numeric(final_df["fuzzy_score"], errors="coerce")

# Build BD_UserID -> Fynancial_BlackDiamondId lookup
bd_userid_to_fyn_bd = {}
for _, row in final_df.iterrows():
    bd_uid = row["BD_UserID"]
    fyn_bd = row["Fynancial_BlackDiamondId"]
    if pd.notna(bd_uid) and pd.notna(fyn_bd) and str(fyn_bd).strip() not in ("", "nan"):
        bd_userid_to_fyn_bd[bd_uid] = str(fyn_bd)

final_df["_bd_userid_sort"] = pd.to_numeric(final_df["BD_UserID"], errors="coerce")

# -----------------------------
# Sort
# -----------------------------
final_df = final_df.sort_values(
    by=["Fynancial_BlackDiamondId", "_bd_userid_sort", "Fynancial_LastName"],
    na_position="last"
).drop(columns=["_bd_userid_sort"]).reset_index(drop=True)

# Post-sort: ensure all rows sharing a BD_UserID are adjacent
processed_bd_ids = set()
new_order = []
used_indices = set()

for i in range(len(final_df)):
    if i in used_indices:
        continue
    bd_id = final_df.at[i, "BD_UserID"]
    if pd.notna(bd_id) and bd_id in duplicate_bd_userids and bd_id not in processed_bd_ids:
        group_indices = final_df[final_df["BD_UserID"] == bd_id].index.tolist()
        new_order.extend(group_indices)
        used_indices.update(group_indices)
        processed_bd_ids.add(bd_id)
    else:
        new_order.append(i)
        used_indices.add(i)

final_df = final_df.loc[new_order].reset_index(drop=True)

# -----------------------------
# Group classification and color assignment
# -----------------------------
final_df["_gk"] = final_df.apply(get_group_key, axis=1)

# Merge group keys: if a Fyn BD ID group contains a member whose BD_UserID is a duplicate,
# redirect all members of that Fyn BD ID group to the bd_ group key
fyn_bd_to_bd_group = {}
for _, row in final_df.iterrows():
    fyn_bd = row["Fynancial_BlackDiamondId"]
    bd_uid = row["BD_UserID"]
    if (pd.notna(fyn_bd) and str(fyn_bd).strip() not in ("", "nan") and
            pd.notna(bd_uid) and bd_uid in duplicate_bd_userids):
        fyn_bd_to_bd_group[str(fyn_bd)] = f"bd_{bd_uid}"

def get_merged_gk(row):
    fyn_bd = row["Fynancial_BlackDiamondId"]
    if pd.notna(fyn_bd) and str(fyn_bd).strip() not in ("", "nan"):
        fyn_bd_str = str(fyn_bd)
        if fyn_bd_str in fyn_bd_to_bd_group:
            return fyn_bd_to_bd_group[fyn_bd_str]
    return get_group_key(row)

final_df["_gk"] = final_df.apply(get_merged_gk, axis=1)

confident_keys = set()
attention_keys = set()
group_color = {}

for key, grp in final_df.groupby("_gk", sort=False):
    # Evaluate match types excluding BD_No_Fynancial_Match rows so they don't
    # drag down the color of an otherwise confident group
    non_bd_types = [mt for mt in grp["match_type"].tolist() if mt != "BD_No_Fynancial_Match"]
    all_types = grp["match_type"].tolist()
    has_shared = any(grp["dupe_bd_userid_found"] == True)
    all_bd_no_fyn = all(mt == "BD_No_Fynancial_Match" for mt in all_types)

    if all_bd_no_fyn:
        attention_keys.add(key)
        group_color[key] = PURPLE
    elif non_bd_types and all(mt == "Email_And_FullName" for mt in non_bd_types) and not has_shared:
        confident_keys.add(key)
        group_color[key] = DARK_GREEN
    elif non_bd_types and all(mt in ("Email", "FullName", "Email_And_FullName") for mt in non_bd_types) and not has_shared:
        confident_keys.add(key)
        group_color[key] = LIGHT_GREEN
    elif has_shared or all(mt == "No_match" for mt in all_types) or any(mt in ("Matched_No_Fyn_BD_ID", "Unmatched_No_Fyn_BD_ID") for mt in all_types):
        fyn_bd_values = grp["Fynancial_BlackDiamondId"].apply(
            lambda v: None if pd.isna(v) or str(v).strip() in ("", "nan") else str(v)
        )
        has_some = fyn_bd_values.notna().any()
        has_missing = fyn_bd_values.isna().any()
        unique_ids = fyn_bd_values.dropna().unique()

        if not has_some:
            attention_keys.add(key)
            group_color[key] = LIGHT_BLUE
        elif has_some and (has_missing or len(unique_ids) > 1):
            attention_keys.add(key)
            group_color[key] = DARK_RED
        else:
            attention_keys.add(key)
            group_color[key] = LIGHT_RED
    else:
        attention_keys.add(key)
        group_color[key] = LIGHT_YELLOW

# Build color lookup keyed by Fynancial_UserUniqueId (and BD_UserID as fallback for BD_No_Fynancial_Match rows)
final_df["_row_color"] = final_df["_gk"].map(group_color)

# Force any group containing FullName, Fuzzy_Match, or Household_Name to light yellow
REVIEW_TYPES = {"FullName", "Fuzzy_Match", "Household_Name"}
for key, grp in final_df.groupby("_gk", sort=False):
    if any(mt in REVIEW_TYPES for mt in grp["match_type"].tolist()):
        if group_color.get(key) in (DARK_GREEN, LIGHT_GREEN):
            group_color[key] = LIGHT_YELLOW
            attention_keys.add(key)
            confident_keys.discard(key)
            for idx in grp.index:
                final_df.at[idx, "_row_color"] = LIGHT_YELLOW

# Force any group containing a BD_No_Fynancial_Match row to DARK_RED in attention,
# regardless of what the other rows in the group look like.
# Tag these groups so they sort to the bottom of the dark reds.
bd_no_fyn_group_keys = set()
for key, grp in final_df.groupby("_gk", sort=False):
    if any(mt == "BD_No_Fynancial_Match" for mt in grp["match_type"].tolist()):
        # Only apply if not a pure BD_No_Fynancial_Match group (those stay PURPLE)
        if not all(mt == "BD_No_Fynancial_Match" for mt in grp["match_type"].tolist()):
            group_color[key] = DARK_RED
            attention_keys.add(key)
            confident_keys.discard(key)
            bd_no_fyn_group_keys.add(key)
            for idx in grp.index:
                final_df.at[idx, "_row_color"] = DARK_RED

color_lookup = {}
for _, row in final_df.iterrows():
    uid = row.get("Fynancial_UserUniqueId")
    color = row.get("_row_color")
    if pd.notna(uid) and str(uid).strip() not in ("", "nan"):
        color_lookup[uid] = color
    # Also key by BD_UserID so BD_No_Fynancial_Match rows (no Fyn UID) can be colored
    bd_uid = row.get("BD_UserID")
    if pd.notna(bd_uid):
        color_lookup.setdefault(f"_bd_{bd_uid}", color)

final_df = final_df.drop(columns=["_gk", "_row_color"])

# -----------------------------
# Split into confident and attention sections
# -----------------------------
def row_in_confident(row):
    key = get_merged_gk(row)
    return key in confident_keys and key not in bd_no_fyn_group_keys

df_confident = final_df[final_df.apply(row_in_confident, axis=1)].copy()
df_attention = final_df[~final_df.apply(row_in_confident, axis=1)].copy()

# Re-run BD_UserID adjacency pass within each section after the split
def group_by_bd_userid(df):
    processed = set()
    new_order = []
    used = set()
    for i in range(len(df)):
        if i in used:
            continue
        bd_id = df.iloc[i]["BD_UserID"]
        if pd.notna(bd_id) and bd_id in duplicate_bd_userids and bd_id not in processed:
            group_indices = df[df["BD_UserID"] == bd_id].index.tolist()
            pos_indices = [df.index.get_loc(gi) for gi in group_indices]
            new_order.extend(pos_indices)
            used.update(pos_indices)
            processed.add(bd_id)
        else:
            new_order.append(i)
            used.add(i)
    return df.iloc[new_order].reset_index(drop=True)

df_confident = group_by_bd_userid(df_confident)
df_attention = group_by_bd_userid(df_attention)

# If dupe_bd_userid_found rows have their partner in a different section, move them to attention
dupe_bd_ids_in_attention = set(df_attention[df_attention["dupe_bd_userid_found"] == True]["BD_UserID"].dropna().unique())
rows_to_move = df_confident[df_confident["BD_UserID"].isin(dupe_bd_ids_in_attention)]
if len(rows_to_move) > 0:
    df_confident = df_confident[~df_confident["BD_UserID"].isin(dupe_bd_ids_in_attention)].reset_index(drop=True)
    df_attention = pd.concat([rows_to_move, df_attention], ignore_index=True)
    df_attention = group_by_bd_userid(df_attention)

# Sort each section by color rank then group size then original order
def sort_by_color(df):
    df = df.copy()
    gk_series = df.apply(get_merged_gk, axis=1)
    group_sizes = gk_series.map(gk_series.value_counts())

    gk_has_match = {}
    for key, grp in df.groupby(gk_series, sort=False):
        gk_has_match[key] = 0 if any(pd.notna(v) for v in grp["BD_UserID"]) else 1
    has_match_series = gk_series.map(gk_has_match)

    gk_has_nomatch = {}
    for key, grp in df.groupby(gk_series, sort=False):
        gk_has_nomatch[key] = 1 if any(mt == "No_match" for mt in grp["match_type"]) else 0
    has_nomatch_series = gk_series.map(gk_has_nomatch)

    def get_row_color(r):
        uid = r.get("Fynancial_UserUniqueId")
        if pd.notna(uid) and str(uid).strip() not in ("", "nan"):
            return color_lookup.get(uid)
        bd_uid = r.get("BD_UserID")
        if pd.notna(bd_uid):
            return color_lookup.get(f"_bd_{bd_uid}")
        return None

    df["_cr"] = df.apply(lambda r: COLOR_RANK.get(get_row_color(r), 99), axis=1)
    df["_group_size"] = group_sizes.values
    df["_has_match"] = has_match_series.values
    df["_has_nomatch"] = has_nomatch_series.values
    df["_orig"] = range(len(df))
    # Within dark red groups, push those containing BD_No_Fynancial_Match to the bottom
    df["_has_bd_no_fyn"] = df.apply(
        lambda r: 1 if get_merged_gk(r) in bd_no_fyn_group_keys else 0, axis=1
    )
    df["_match_sort"] = df.apply(
        lambda r: r["_has_match"] if get_row_color(r) in (LIGHT_RED, DARK_RED) else 0,
        axis=1
    )
    df["_yellow_nomatch_sort"] = df.apply(
        lambda r: r["_has_nomatch"] if (
            group_sizes[r.name] > 1 and get_row_color(r) == LIGHT_YELLOW
        ) else 0,
        axis=1
    )
    match_type_order = {
        "Email_And_FullName": 0, "Email": 1, "FullName": 2,
        "Fuzzy_Match": 3, "Household_Email": 4, "Household_Name": 5,
        "Household_Partial": 6, "Duplicate_BD_UserID_Match": 7,
        "Matched_No_Fyn_BD_ID": 8, "Unmatched_No_Fyn_BD_ID": 9,
        "No_match": 10, "BD_No_Fynancial_Match": 11,
    }
    df["_fuzzy_sort"] = df.apply(
        lambda r: -r["fuzzy_score"] if (
            group_sizes[r.name] == 1 and
            get_row_color(r) in (LIGHT_GREEN, LIGHT_YELLOW) and
            pd.notna(r["fuzzy_score"])
        ) else 0,
        axis=1
    )
    df["_mt_sort"] = df.apply(
        lambda r: match_type_order.get(r["match_type"], 99) if (
            group_sizes[r.name] == 1 and
            get_row_color(r) in (LIGHT_GREEN, LIGHT_YELLOW)
        ) else 0,
        axis=1
    )
    df = df.sort_values(
        ["_cr", "_has_bd_no_fyn", "_group_size", "_match_sort", "_yellow_nomatch_sort", "_mt_sort", "_fuzzy_sort", "_orig"],
        ascending=[True, True, False, True, True, True, True, True]
    ).drop(columns=["_cr", "_has_bd_no_fyn", "_group_size", "_has_match", "_has_nomatch", "_match_sort",
                    "_yellow_nomatch_sort", "_mt_sort", "_fuzzy_sort", "_orig"])
    return df.reset_index(drop=True)

df_confident = sort_by_color(df_confident)
df_attention = sort_by_color(df_attention)

# -----------------------------
# Insert blank rows between groups within each section
# -----------------------------
def insert_blanks(df):
    blank = pd.Series({col: None for col in df.columns})
    rows = []
    prev_key = "SENTINEL"
    for i, row in df.iterrows():
        curr_key = get_merged_gk(row)
        if i > 0 and curr_key != prev_key:
            rows.append(blank)
        rows.append(row)
        prev_key = curr_key
    return pd.DataFrame(rows).reset_index(drop=True)

excel_confident = insert_blanks(df_confident)
excel_attention = insert_blanks(df_attention)
section_blank = pd.DataFrame([{col: None for col in final_df.columns}] * 2)
final_df_excel = pd.concat([excel_confident, section_blank, excel_attention], ignore_index=True)

# Drop internal column before writing
final_df_excel = final_df_excel.drop(columns=["_forced_group_bd_uid"], errors="ignore")
final_df = final_df.drop(columns=["_forced_group_bd_uid"], errors="ignore")

# -----------------------------
# Write Excel
# -----------------------------
output_filename = os.path.join(FIRM_DIR, f"{FIRM_NAME}_fynancial_blackdiamond_mapping.xlsx")
final_df_excel.to_excel(output_filename, index=False)

wb = load_workbook(output_filename)
ws = wb.active

# Auto-size columns
for col_idx, col in enumerate(final_df_excel.columns, start=1):
    max_len = max(len(str(col)), final_df_excel[col].astype(str).str.len().max())
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

# Column dividers
divider_side = Side(style="medium")
divider_cols = [
    final_df_excel.columns.get_loc("Fynancial_BlackDiamondId") + 1,
    final_df_excel.columns.get_loc("BD_LastName") + 1,
]

# Bold header
for cell in ws[1]:
    cell.font = Font(bold=True)

# Color rows
match_type_col = final_df_excel.columns.get_loc("match_type") + 1
shared_col = final_df_excel.columns.get_loc("dupe_bd_userid_found") + 1
uid_col = final_df_excel.columns.get_loc("Fynancial_UserUniqueId") + 1
bd_uid_col = final_df_excel.columns.get_loc("BD_UserID") + 1

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    uid = row[uid_col - 1].value
    bd_uid = row[bd_uid_col - 1].value
    match_val = row[match_type_col - 1].value

    # Determine color: prefer Fyn UID lookup, fall back to BD UID lookup
    if uid and uid in color_lookup:
        hex_color = color_lookup[uid]
    elif bd_uid and f"_bd_{bd_uid}" in color_lookup:
        hex_color = color_lookup[f"_bd_{bd_uid}"]
    else:
        hex_color = None

    for cell in row:
        if hex_color:
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        if cell.column in divider_cols:
            existing = cell.border
            cell.border = Border(
                left=existing.left,
                right=divider_side,
                top=existing.top,
                bottom=existing.bottom
            )

    shared_cell = row[shared_col - 1]
    if shared_cell.value is True:
        row_color = color_lookup.get(uid) or color_lookup.get(f"_bd_{bd_uid}", LIGHT_RED)
        if row_color == DARK_RED:
            dupe_color = DARK_RED
        elif row_color == LIGHT_BLUE:
            dupe_color = LIGHT_BLUE
        else:
            dupe_color = LIGHT_RED
        shared_cell.fill = PatternFill(start_color=dupe_color, end_color=dupe_color, fill_type="solid")

# Merge notes cells per group
notes_col = final_df_excel.columns.get_loc("notes") + 1

def get_excel_group_key(row):
    bd = row.get("Fynancial_BlackDiamondId")
    bd_uid = row.get("BD_UserID")
    uid = row.get("Fynancial_UserUniqueId")
    if pd.notna(bd) and str(bd).strip() not in ("", "nan"):
        return str(bd)
    if pd.notna(bd_uid) and bd_uid in bd_userid_to_fyn_bd:
        return bd_userid_to_fyn_bd[bd_uid]
    if pd.notna(bd_uid) and bd_uid in duplicate_bd_userids:
        return f"bd_{bd_uid}"
    return f"single_{uid}"

rows_with_keys = []
for i, row in final_df_excel.iterrows():
    if all(pd.isna(v) or v is None for v in row.values):
        rows_with_keys.append((i, None))
    else:
        rows_with_keys.append((i, get_excel_group_key(row)))

i = 0
while i < len(rows_with_keys):
    df_idx, key = rows_with_keys[i]
    if key is None:
        i += 1
        continue
    j = i + 1
    while j < len(rows_with_keys) and rows_with_keys[j][1] == key:
        j += 1
    excel_start = df_idx + 2
    excel_end = rows_with_keys[j - 1][0] + 2
    if excel_end > excel_start:
        ws.merge_cells(
            start_row=excel_start, start_column=notes_col,
            end_row=excel_end, end_column=notes_col
        )
    merged_cell = ws.cell(row=excel_start, column=notes_col)
    merged_cell.border = Border(
        top=Side(style="thin"), bottom=Side(style="thin"),
        left=Side(style="thin"), right=Side(style="thin")
    )
    i = j

wb.save(output_filename)

# -----------------------------
# Draw thick border boxes around groups of 2+ rows
# -----------------------------
wb = load_workbook(output_filename)
ws = wb.active

thick = Side(style="medium", color="000000")

row_group_keys = []
for i, row in final_df_excel.iterrows():
    if all(pd.isna(v) or v is None for v in row.values):
        row_group_keys.append((i + 2, None))
    else:
        row_group_keys.append((i + 2, get_excel_group_key(row)))

i = 0
while i < len(row_group_keys):
    excel_row, key = row_group_keys[i]
    if key is None:
        i += 1
        continue
    j = i + 1
    while j < len(row_group_keys) and row_group_keys[j][1] == key:
        j += 1
    group_excel_rows = [r for r, k in row_group_keys[i:j] if k is not None]
    if len(group_excel_rows) > 1:
        first_row = min(group_excel_rows)
        last_row = max(group_excel_rows)
        max_col = len(final_df_excel.columns)
        for excel_row in group_excel_rows:
            for col in range(1, max_col + 1):
                cell = ws.cell(row=excel_row, column=col)
                top = thick if excel_row == first_row else cell.border.top
                bottom = thick if excel_row == last_row else cell.border.bottom
                left = thick if col == 1 else cell.border.left
                right = thick if col == max_col else cell.border.right
                if col in divider_cols:
                    right = thick if col == max_col else Side(style="medium")
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)
    i = j

wb.save(output_filename)

# -----------------------------
# Also write CSV
# -----------------------------
csv_filename = os.path.join(FIRM_DIR, f"{FIRM_NAME}_fynancial_blackdiamond_mapping.csv")
final_df.to_csv(csv_filename, index=False)

fyn_row_count = len(fyn_df)
bd_unmatched_count = len(final_df[final_df["match_type"] == "BD_No_Fynancial_Match"])
assert len(final_df) == fyn_row_count + bd_unmatched_count, "Row count mismatch!"

print(f"✅ Excel generated: {output_filename}")
print(f"✅ CSV generated:   {csv_filename}")
print(f"\nSummary:")
print(f"  Total records:            {len(final_df)}")
print(f"  Email And FullName:       {len(final_df[final_df['match_type'] == 'Email_And_FullName'])}")
print(f"  Email:                    {len(final_df[final_df['match_type'] == 'Email'])}")
print(f"  FullName:                 {len(final_df[final_df['match_type'] == 'FullName'])}")
print(f"  Fuzzy Match:              {len(final_df[final_df['match_type'] == 'Fuzzy_Match'])}")
print(f"  Household Email:          {len(final_df[final_df['match_type'] == 'Household_Email'])}")
print(f"  Household Name:           {len(final_df[final_df['match_type'] == 'Household_Name'])}")
print(f"  Household Partial:        {len(final_df[final_df['match_type'] == 'Household_Partial'])}")
print(f"  Duplicate BD UserID:      {len(final_df[final_df['match_type'] == 'Duplicate_BD_UserID_Match'])}")
print(f"  Matched No Fyn BD ID:     {len(final_df[final_df['match_type'] == 'Matched_No_Fyn_BD_ID'])}")
print(f"  Unmatched No Fyn BD ID:   {len(final_df[final_df['match_type'] == 'Unmatched_No_Fyn_BD_ID'])}")
print(f"  No match:                 {len(final_df[final_df['match_type'] == 'No_match'])}")
print(f"  BD no Fynancial:          {len(final_df[final_df['match_type'] == 'BD_No_Fynancial_Match'])}")